"""
Togal Takeoff Converter – Local Web App
Run:  python app.py
Then open:  http://localhost:5000
"""

import os, re, sys
from flask import Flask, render_template, request, jsonify, send_from_directory
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY      = "1F3864"
WHITE_FG  = "FFFFFF"
AMBER     = "F4B942"
BLACK     = "000000"
GREY      = "D9D9D9"
PALE_YEL  = "FFF2CC"
WHITE     = "FFFFFF"
DOM_BLUE  = "BDD7EE"
IMP_GREEN = "C6EFCE"
DIFF_PUR  = "7030A0"
DIFF_BG   = "EAD1F5"

THIN = Side(style="thin", color="AAAAAA")
def tb(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sc(cell, bold=False, bg=None, fg=BLACK, align="left",
       size=10, wrap=False, fmt=None, italic=False):
    cell.font = Font(name="Calibri", bold=bold, italic=italic, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg, fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = tb()
    if fmt:
        cell.number_format = fmt

# ── Parse Togal export ────────────────────────────────────────────────────────
SKIP_CATS = {"togal counts", "gross sqft"}

def parse_togal(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    categories = []
    current_cat = None

    for row in rows[1:]:
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None

        a_str = str(a).strip() if a else ""
        c_str = str(c).strip() if c else ""

        if not a_str:
            continue
        if a_str.lower() == "total":
            current_cat = None
            continue
        if b is None:
            if a_str.lower() in SKIP_CATS:
                current_cat = None
                continue
            current_cat = {"name": a_str, "items": []}
            categories.append(current_cat)
            continue
        if current_cat is not None:
            try:
                qty = float(b) if b is not None else 0
            except (ValueError, TypeError):
                qty = 0
            current_cat["items"].append({
                "item": a_str,
                "qty":  qty,
                "uom":  c_str if c_str else "EA",
            })

    return categories

# ── Build Price List tab ──────────────────────────────────────────────────────
def build_price_list(wb, categories):
    ws = wb.create_sheet("Price List")
    for i, w in enumerate([36, 18, 16, 10, 36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for c, h in enumerate(["Classification / Item","Part Number",
                            "Unit Price (Domestic)","UOM","Notes / Vendor"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        sc(cell, bold=True, bg=NAVY, fg=WHITE_FG, align="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    seen = {}
    pr = 2
    for cat in categories:
        ws.merge_cells(f"A{pr}:E{pr}")
        sc(ws.cell(row=pr, column=1, value=cat["name"].upper()),
           bold=True, bg=AMBER, align="left")
        ws.row_dimensions[pr].height = 16
        pr += 1

        for item in cat["items"]:
            key = item["item"]
            if key in seen:
                continue
            seen[key] = pr
            ws.cell(row=pr, column=1, value=key)
            ws.cell(row=pr, column=2, value="")
            ws.cell(row=pr, column=3, value="")
            ws.cell(row=pr, column=4, value=item["uom"])
            ws.cell(row=pr, column=5, value="")
            bg = PALE_YEL if pr % 2 == 0 else WHITE
            for col in range(1, 6):
                cell = ws.cell(row=pr, column=col)
                sc(cell, bg=bg, fmt='"$"#,##0.00' if col == 3 else None)
            ws.row_dimensions[pr].height = 16
            pr += 1

    note = ws.cell(row=pr+1, column=1,
                   value="⬆  Fill in Part Number, Unit Price (Domestic), and verify UOM for each item.")
    note.font = Font(name="Calibri", italic=True, color="666666", size=9)
    return ws, seen

# ── Build Dom vs Import tab ───────────────────────────────────────────────────
def build_dom_import(wb, categories, pl_seen):
    ws = wb.create_sheet("Dom vs Import")
    for i, w in enumerate([36,12,8,16,16,16,16,16,18,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    def grp(rng, label, bg):
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        sc(cell, bold=True, bg=bg, fg=WHITE_FG, align="center", size=10)

    grp("A1:C1", "ITEM",                            NAVY)
    grp("D1:E1", "🏠  DOMESTIC",                    "2E75B6")
    grp("F1:G1", "🚢  IMPORT",                      "375623")
    grp("H1:J1", "📊  DIFFERENCE  (Import vs Dom)", DIFF_PUR)
    ws.row_dimensions[1].height = 16

    for col, label, bg, fg in [
        (1,"Classification",  NAVY,     WHITE_FG),
        (2,"Qty",             NAVY,     WHITE_FG),
        (3,"UOM",             NAVY,     WHITE_FG),
        (4,"Dom Unit $",      "2E75B6", WHITE_FG),
        (5,"Dom Total",       "2E75B6", WHITE_FG),
        (6,"Import Unit $",   "375623", WHITE_FG),
        (7,"Import Total",    "375623", WHITE_FG),
        (8,"$ Diff/Unit",     DIFF_PUR, WHITE_FG),
        (9,"Total $ Savings", DIFF_PUR, WHITE_FG),
        (10,"% Savings",      DIFF_PUR, WHITE_FG),
    ]:
        cell = ws.cell(row=2, column=col, value=label)
        sc(cell, bold=True, bg=bg, fg=fg, align="center", size=9)
    ws.row_dimensions[2].height = 20

    dr = 3
    all_data_rows = []

    for cat in categories:
        ws.merge_cells(f"A{dr}:J{dr}")
        sc(ws.cell(row=dr, column=1, value=cat["name"].upper()),
           bold=True, bg=AMBER, align="left")
        ws.row_dimensions[dr].height = 16
        dr += 1
        cat_start = dr

        for item in cat["items"]:
            bg = PALE_YEL if dr % 2 == 0 else WHITE
            pl_row = pl_seen.get(item["item"])
            qty = item["qty"]
            fmt_qty = "#,##0" if qty == int(qty) else "#,##0.00"

            sc(ws.cell(row=dr, column=1, value=item["item"]), bg=bg)
            c = ws.cell(row=dr, column=2, value=qty)
            sc(c, bg=bg, align="right", fmt=fmt_qty)
            if pl_row:
                c = ws.cell(row=dr, column=3, value=f"='Price List'!D{pl_row}")
            else:
                c = ws.cell(row=dr, column=3, value=item["uom"])
            sc(c, bg=bg, align="center")
            if pl_row:
                c = ws.cell(row=dr, column=4, value=f"='Price List'!C{pl_row}")
            else:
                c = ws.cell(row=dr, column=4, value="")
            sc(c, bg=DOM_BLUE, fmt='"$"#,##0.00')
            c = ws.cell(row=dr, column=5, value=f"=B{dr}*D{dr}")
            sc(c, bg=DOM_BLUE, align="right", bold=True, fmt='"$"#,##0.00')
            c = ws.cell(row=dr, column=6, value="")
            sc(c, bg=IMP_GREEN, fmt='"$"#,##0.00')
            c = ws.cell(row=dr, column=7,
                        value=f"=IF(F{dr}=\"\",\"\",B{dr}*F{dr})")
            sc(c, bg=IMP_GREEN, align="right", bold=True, fmt='"$"#,##0.00')
            c = ws.cell(row=dr, column=8,
                        value=f"=IF(F{dr}=\"\",\"\",F{dr}-D{dr})")
            sc(c, bg=DIFF_BG, align="right",
               fmt='"$"#,##0.00;[Red]"($"#,##0.00")"')
            c = ws.cell(row=dr, column=9,
                        value=f"=IF(F{dr}=\"\",\"\",E{dr}-G{dr})")
            sc(c, bg=DIFF_BG, align="right", bold=True,
               fmt='"$"#,##0.00;[Red]"($"#,##0.00")"')
            c = ws.cell(row=dr, column=10,
                        value=f"=IF(OR(F{dr}=\"\",D{dr}=0),\"\",(D{dr}-F{dr})/D{dr})")
            sc(c, bg=DIFF_BG, align="right", fmt='0.0%;[Red](0.0%)')

            ws.row_dimensions[dr].height = 16
            all_data_rows.append(dr)
            dr += 1

        cat_end = dr - 1
        sc(ws.cell(row=dr, column=1, value="Subtotal"), bold=True, bg=GREY)
        for col in range(2, 11):
            sc(ws.cell(row=dr, column=col, value=""), bg=GREY)
        c = ws.cell(row=dr, column=5, value=f"=SUM(E{cat_start}:E{cat_end})")
        sc(c, bold=True, bg=GREY, align="right", fmt='"$"#,##0.00')
        c = ws.cell(row=dr, column=7,
                    value=f"=SUMIF(F{cat_start}:F{cat_end},\"<>\"\"\",G{cat_start}:G{cat_end})")
        sc(c, bold=True, bg=GREY, align="right", fmt='"$"#,##0.00')
        c = ws.cell(row=dr, column=9,
                    value=f"=SUMIF(I{cat_start}:I{cat_end},\"<>\"\"\",I{cat_start}:I{cat_end})")
        sc(c, bold=True, bg=GREY, align="right",
           fmt='"$"#,##0.00;[Red]"($"#,##0.00")"')
        ws.row_dimensions[dr].height = 18
        dr += 2

    if not all_data_rows:
        return ws

    first, last = all_data_rows[0], all_data_rows[-1]
    gt = dr
    ws.merge_cells(f"A{gt}:D{gt}")
    sc(ws.cell(row=gt, column=1, value="GRAND TOTAL — DOMESTIC"),
       bold=True, bg=NAVY, fg=WHITE_FG, align="right", size=11)
    c = ws.cell(row=gt, column=5,
                value=f"=SUMPRODUCT((D{first}:D{last}<>\"\")*B{first}:B{last}*D{first}:D{last})")
    sc(c, bold=True, bg="2E75B6", fg=WHITE_FG, align="right", size=11, fmt='"$"#,##0.00')
    ws.merge_cells(f"F{gt}:H{gt}")
    sc(ws.cell(row=gt, column=6, value="GRAND TOTAL — IMPORT"),
       bold=True, bg="375623", fg=WHITE_FG, align="right", size=11)
    c = ws.cell(row=gt, column=9,
                value=f"=SUMPRODUCT((F{first}:F{last}<>\"\")*B{first}:B{last}*F{first}:F{last})")
    sc(c, bold=True, bg="375623", fg=WHITE_FG, align="right", size=11, fmt='"$"#,##0.00')
    ws.row_dimensions[gt].height = 24

    ns = gt + 1
    ws.merge_cells(f"A{ns}:H{ns}")
    sc(ws.cell(row=ns, column=1, value="NET SAVINGS  (Domestic − Import)"),
       bold=True, bg=DIFF_PUR, fg=WHITE_FG, align="right", size=11)
    c = ws.cell(row=ns, column=9, value=f"=E{gt}-I{gt}")
    sc(c, bold=True, bg=DIFF_PUR, fg=WHITE_FG, align="right", size=11,
       fmt='"$"#,##0.00;[Red]"($"#,##0.00")"')
    c = ws.cell(row=ns, column=10, value=f"=IF(E{gt}=0,\"\",I{ns}/E{gt})")
    sc(c, bold=True, bg=DIFF_PUR, fg=WHITE_FG, align="right", size=11,
       fmt='0.0%;[Red](0.0%)')
    ws.row_dimensions[ns].height = 24

    inst = ws.cell(row=ns+2, column=1,
                   value=("ℹ  Instructions: Domestic Unit $ (col D) auto-fills from Price List. "
                          "Enter Import Unit $ in col F. Positive savings = import is cheaper. "
                          "Red = import costs more."))
    ws.merge_cells(f"A{ns+2}:J{ns+2}")
    inst.font = Font(name="Calibri", italic=True, color="555555", size=9)
    inst.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ns+2].height = 28

    return ws

# ── Build Takeoff tab ─────────────────────────────────────────────────────────
def build_takeoff(wb, categories, pl_seen):
    ws = wb.create_sheet("Takeoff", 0)
    for col, w in zip("ABCDE", [36, 14, 14, 44, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    for c, h in enumerate(["Classification","Quantity 1","Quantity1 UOM",
                            "Part # / Description","Extended Total"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        sc(cell, bold=True, bg=NAVY, fg=WHITE_FG, align="center")
    ws.row_dimensions[1].height = 22

    cur = 2
    cat_total_rows = []

    for cat in categories:
        ws.merge_cells(f"A{cur}:E{cur}")
        sc(ws.cell(row=cur, column=1, value=cat["name"].upper()),
           bold=True, bg=AMBER, align="left")
        ws.row_dimensions[cur].height = 18
        cur += 1
        data_start = cur

        for i, item in enumerate(cat["items"]):
            bg = PALE_YEL if i % 2 == 0 else WHITE
            r = cur
            pl_row = pl_seen.get(item["item"])
            qty = item["qty"]
            fmt_qty = "#,##0" if qty == int(qty) else "#,##0.00"

            sc(ws.cell(row=r, column=1, value=item["item"]), bg=bg)
            c = ws.cell(row=r, column=2, value=qty)
            sc(c, bg=bg, align="right", fmt=fmt_qty)
            if pl_row:
                c = ws.cell(row=r, column=3, value=f"='Price List'!D{pl_row}")
            else:
                c = ws.cell(row=r, column=3, value=item["uom"])
            sc(c, bg=bg, align="center")
            if pl_row:
                c = ws.cell(row=r, column=4,
                            value=(f"='Price List'!B{pl_row}"
                                   f"&IF('Price List'!B{pl_row}<>\"\",\"  –  \",\"\")"
                                   f"&IF('Price List'!C{pl_row}<>\"\","
                                   f"\"$\"&TEXT('Price List'!C{pl_row},\"#,##0.00\")"
                                   f"&\" \"&'Price List'!D{pl_row}&\" EACH\","
                                   f"\"— enter price in Price List tab —\")"))
            else:
                c = ws.cell(row=r, column=4, value="— enter in Price List tab —")
            sc(c, bg=bg, wrap=True)
            if pl_row:
                c = ws.cell(row=r, column=5,
                            value=f"=IF('Price List'!C{pl_row}=\"\",\"\",B{r}*'Price List'!C{pl_row})")
                sc(c, bg=bg, align="right", fmt='"$"#,##0.00')
            else:
                sc(ws.cell(row=r, column=5, value=""), bg=bg)

            ws.row_dimensions[r].height = 16
            cur += 1

        data_end = cur - 1
        sc(ws.cell(row=cur, column=1, value="Total"), bold=True, bg=GREY)
        c = ws.cell(row=cur, column=2, value=f"=SUM(B{data_start}:B{data_end})")
        sc(c, bold=True, bg=GREY, align="right", fmt="#,##0.00")
        sc(ws.cell(row=cur, column=3, value=""), bg=GREY)
        sc(ws.cell(row=cur, column=4, value="SECTION TOTAL"),
           bold=True, bg=GREY, align="center")
        c = ws.cell(row=cur, column=5,
                    value=f"=SUMIF(E{data_start}:E{data_end},\"<>\"\"\",E{data_start}:E{data_end})")
        sc(c, bold=True, bg=GREY, align="right", fmt='"$"#,##0.00')
        ws.row_dimensions[cur].height = 18
        cat_total_rows.append(cur)
        cur += 2

    cur += 1
    ws.merge_cells(f"A{cur}:D{cur}")
    sc(ws.cell(row=cur, column=1, value="GRAND TOTAL"),
       bold=True, bg=NAVY, fg=WHITE_FG, align="right", size=11)
    refs = "+".join([f"E{r}" for r in cat_total_rows])
    c = ws.cell(row=cur, column=5, value=f"={refs}")
    sc(c, bold=True, bg=NAVY, fg=WHITE_FG, align="right", size=11, fmt='"$"#,##0.00')
    ws.row_dimensions[cur].height = 22
    return ws

# ── Flask routes ──────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/convert", methods=["POST"])
def convert():
    if "file" not in request.files:
        return jsonify(error="No file uploaded."), 400

    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify(error="Please upload a .xlsx file."), 400

    project_name = request.form.get("project_name", "").strip()
    if not project_name:
        project_name = f.filename.replace(".xlsx", "")

    # Save upload to temp
    upload_path = os.path.join(OUTPUT_DIR, "upload_tmp.xlsx")
    f.save(upload_path)

    try:
        categories = parse_togal(upload_path)
    except Exception as e:
        return jsonify(error=str(e)), 400

    if not categories:
        return jsonify(error="No recognisable data found in the Togal file. "
                             "Make sure it uses the standard Togal export format."), 400

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)
    _, pl_seen = build_price_list(wb, categories)
    build_dom_import(wb, categories, pl_seen)
    build_takeoff(wb, categories, pl_seen)

    # Safe filename
    safe = re.sub(r'[^\w\s\-]', '', project_name).strip()
    safe = re.sub(r'\s+', '_', safe)[:60]
    out_filename = f"{safe}_Takeoff.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)

    total_items = sum(len(c["items"]) for c in categories)
    return jsonify(
        filename=out_filename,
        categories=len(categories),
        items=total_items
    )

@app.route("/download/<filename>")
def download(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)

# ── Launch ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = 5000
    print(f"\n✅  Togal Takeoff Converter running at http://127.0.0.1:{port}")
    print("   Close this window to stop the app.\n")
    app.run(host="127.0.0.1", port=port, debug=False)
